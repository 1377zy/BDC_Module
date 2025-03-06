from flask import render_template, flash, redirect, url_for, request, jsonify, current_app, send_file
from flask_login import login_required, current_user
from app import db
from app.analytics import bp
from datetime import datetime, timedelta
import json
import calendar
import io
import xlsxwriter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from app.models_main import Lead, Appointment, Communication, User, MarketingCampaign, CampaignChannel, CampaignMetric
import csv
import random
from app.decorators import manager_required

@bp.route('/dashboard')
@login_required
def dashboard():
    """Main analytics dashboard with key performance indicators"""
    # Date ranges
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    
    # Lead statistics
    total_leads = Lead.query.count()
    new_leads_today = Lead.query.filter(func.date(Lead.created_at) == today).count()
    new_leads_week = Lead.query.filter(Lead.created_at >= start_of_week).count()
    new_leads_month = Lead.query.filter(Lead.created_at >= start_of_month).count()
    
    # Lead statuses
    lead_statuses = db.session.query(
        Lead.status, func.count(Lead.id)
    ).group_by(Lead.status).all()
    
    # Appointment statistics
    total_appointments = Appointment.query.count()
    appointments_today = Appointment.query.filter(
        func.date(Appointment.date) == today
    ).count()
    appointments_week = Appointment.query.filter(
        Appointment.date >= start_of_week
    ).count()
    appointments_month = Appointment.query.filter(
        Appointment.date >= start_of_month
    ).count()
    
    # Appointment statuses
    appointment_statuses = db.session.query(
        Appointment.status, func.count(Appointment.id)
    ).group_by(Appointment.status).all()
    
    # Communication statistics
    total_communications = Communication.query.count()
    communications_today = Communication.query.filter(
        func.date(Communication.sent_at) == today
    ).count()
    communications_week = Communication.query.filter(
        Communication.sent_at >= start_of_week
    ).count()
    communications_month = Communication.query.filter(
        Communication.sent_at >= start_of_month
    ).count()
    
    # Communication types
    communication_types = db.session.query(
        Communication.type, func.count(Communication.id)
    ).group_by(Communication.type).all()
    
    # User performance (for managers/admins)
    user_performance = None
    if current_user.role in ['admin', 'manager']:
        user_performance = db.session.query(
            User.username,
            func.count(Lead.id).label('leads'),
            func.count(Appointment.id).label('appointments'),
            func.count(Communication.id).label('communications')
        ).outerjoin(Lead, User.id == Lead.user_id
        ).outerjoin(Appointment, User.id == Appointment.user_id
        ).outerjoin(Communication, User.id == Communication.user_id
        ).group_by(User.username).all()
    
    return render_template('analytics/dashboard.html',
                          title='Analytics Dashboard',
                          total_leads=total_leads,
                          new_leads_today=new_leads_today,
                          new_leads_week=new_leads_week,
                          new_leads_month=new_leads_month,
                          lead_statuses=lead_statuses,
                          total_appointments=total_appointments,
                          appointments_today=appointments_today,
                          appointments_week=appointments_week,
                          appointments_month=appointments_month,
                          appointment_statuses=appointment_statuses,
                          total_communications=total_communications,
                          communications_today=communications_today,
                          communications_week=communications_week,
                          communications_month=communications_month,
                          communication_types=communication_types,
                          user_performance=user_performance)

@bp.route('/leads')
@login_required
def lead_analytics():
    """Detailed lead analytics"""
    # Lead source analysis
    lead_sources = db.session.query(
        Lead.source, func.count(Lead.id)
    ).group_by(Lead.source).all()
    
    # Lead status by source
    lead_status_by_source = db.session.query(
        Lead.source, Lead.status, func.count(Lead.id)
    ).group_by(Lead.source, Lead.status).all()
    
    # Lead conversion rates (leads with appointments)
    lead_with_appointments = db.session.query(
        Lead.id
    ).join(Appointment).distinct().count()
    
    conversion_rate = 0
    if total_leads := Lead.query.count():
        conversion_rate = (lead_with_appointments / total_leads) * 100
    
    # Lead age analysis
    current_time = datetime.now()
    lead_age_buckets = {
        'Less than 1 day': 0,
        '1-3 days': 0,
        '4-7 days': 0,
        '1-2 weeks': 0,
        '2-4 weeks': 0,
        'Over 1 month': 0
    }
    
    for lead in Lead.query.all():
        age = (current_time - lead.created_at).days
        if age < 1:
            lead_age_buckets['Less than 1 day'] += 1
        elif 1 <= age <= 3:
            lead_age_buckets['1-3 days'] += 1
        elif 4 <= age <= 7:
            lead_age_buckets['4-7 days'] += 1
        elif 8 <= age <= 14:
            lead_age_buckets['1-2 weeks'] += 1
        elif 15 <= age <= 30:
            lead_age_buckets['2-4 weeks'] += 1
        else:
            lead_age_buckets['Over 1 month'] += 1
    
    # Monthly lead trends
    year = datetime.now().year
    monthly_leads = []
    
    for month in range(1, 13):
        count = Lead.query.filter(
            extract('year', Lead.created_at) == year,
            extract('month', Lead.created_at) == month
        ).count()
        monthly_leads.append({
            'month': calendar.month_name[month],
            'count': count
        })
    
    return render_template('analytics/leads.html',
                          title='Lead Analytics',
                          lead_sources=lead_sources,
                          lead_status_by_source=lead_status_by_source,
                          lead_with_appointments=lead_with_appointments,
                          conversion_rate=conversion_rate,
                          lead_age_buckets=lead_age_buckets,
                          monthly_leads=monthly_leads)

@bp.route('/appointments')
@login_required
def appointment_analytics():
    """Detailed appointment analytics"""
    # Appointment outcome analysis
    appointment_outcomes = db.session.query(
        Appointment.status, func.count(Appointment.id)
    ).group_by(Appointment.status).all()
    
    # Appointment purpose analysis
    appointment_purposes = db.session.query(
        Appointment.purpose, func.count(Appointment.id)
    ).group_by(Appointment.purpose).all()
    
    # Appointment day of week distribution
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_counts = [0] * 7
    
    for appointment in Appointment.query.all():
        day_of_week = appointment.date.weekday()
        day_counts[day_of_week] += 1
    
    day_distribution = [{'day': days[i], 'count': day_counts[i]} for i in range(7)]
    
    # Appointment time of day distribution
    hour_counts = [0] * 24
    
    for appointment in Appointment.query.all():
        hour = appointment.time.hour
        hour_counts[hour] += 1
    
    time_distribution = [{'hour': hour, 'count': hour_counts[hour]} for hour in range(24)]
    
    # No-show rate over time
    year = datetime.now().year
    monthly_no_shows = []
    
    for month in range(1, 13):
        total = Appointment.query.filter(
            extract('year', Appointment.date) == year,
            extract('month', Appointment.date) == month
        ).count()
        
        no_shows = Appointment.query.filter(
            extract('year', Appointment.date) == year,
            extract('month', Appointment.date) == month,
            Appointment.status == 'No-Show'
        ).count()
        
        rate = 0 if total == 0 else (no_shows / total) * 100
        
        monthly_no_shows.append({
            'month': calendar.month_name[month],
            'rate': rate
        })
    
    return render_template('analytics/appointments.html',
                          title='Appointment Analytics',
                          appointment_outcomes=appointment_outcomes,
                          appointment_purposes=appointment_purposes,
                          day_distribution=day_distribution,
                          time_distribution=time_distribution,
                          monthly_no_shows=monthly_no_shows)

@bp.route('/communications')
@login_required
def communication_analytics():
    """Detailed communication analytics"""
    # Communication type distribution
    communication_types = db.session.query(
        Communication.type, func.count(Communication.id)
    ).group_by(Communication.type).all()
    
    # Communication status distribution
    communication_statuses = db.session.query(
        Communication.status, func.count(Communication.id)
    ).group_by(Communication.status).all()
    
    # Communications by hour of day
    hour_counts = [0] * 24
    
    for comm in Communication.query.all():
        hour = comm.sent_at.hour
        hour_counts[hour] += 1
    
    hour_distribution = [{'hour': hour, 'count': hour_counts[hour]} for hour in range(24)]
    
    # Communications by day of week
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_counts = [0] * 7
    
    for comm in Communication.query.all():
        day_of_week = comm.sent_at.weekday()
        day_counts[day_of_week] += 1
    
    day_distribution = [{'day': days[i], 'count': day_counts[i]} for i in range(7)]
    
    # Monthly communication trends
    year = datetime.now().year
    monthly_communications = []
    
    for month in range(1, 13):
        email_count = Communication.query.filter(
            extract('year', Communication.sent_at) == year,
            extract('month', Communication.sent_at) == month,
            Communication.type == 'Email'
        ).count()
        
        sms_count = Communication.query.filter(
            extract('year', Communication.sent_at) == year,
            extract('month', Communication.sent_at) == month,
            Communication.type == 'SMS'
        ).count()
        
        call_count = Communication.query.filter(
            extract('year', Communication.sent_at) == year,
            extract('month', Communication.sent_at) == month,
            Communication.type == 'Call'
        ).count()
        
        monthly_communications.append({
            'month': calendar.month_name[month],
            'email': email_count,
            'sms': sms_count,
            'call': call_count
        })
    
    return render_template('analytics/communications.html',
                          title='Communication Analytics',
                          communication_types=communication_types,
                          communication_statuses=communication_statuses,
                          hour_distribution=hour_distribution,
                          day_distribution=day_distribution,
                          monthly_communications=monthly_communications)

@bp.route('/export/<report_type>/<format>')
@login_required
def export_report(report_type, format):
    """Export analytics reports in CSV or Excel format"""
    if format not in ['csv', 'excel']:
        flash('Invalid export format', 'danger')
        return redirect(url_for('analytics.dashboard'))
    
    # Get date range from request parameters or use default (last 30 days)
    end_date = request.args.get('end_date', datetime.now().date().isoformat())
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).date().isoformat())
    
    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    except ValueError:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
    
    # Create a BytesIO object to store the file
    output = BytesIO()
    
    if report_type == 'performance':
        # Get user performance data
        user_performance = []
        users = User.query.all()
        
        for user in users:
            user_leads = Lead.query.filter(
                Lead.user_id == user.id,
                Lead.created_at >= start_date,
                Lead.created_at <= end_date
            ).count()
            
            user_appointments = Appointment.query.filter(
                Appointment.user_id == user.id,
                Appointment.date >= start_date,
                Appointment.date <= end_date
            ).count()
            
            user_kept_appointments = Appointment.query.filter(
                Appointment.user_id == user.id,
                Appointment.date >= start_date,
                Appointment.date <= end_date,
                Appointment.status == 'completed'
            ).count()
            
            # Count leads that resulted in appointments (conversions)
            user_conversions = db.session.query(Lead.id).join(
                Appointment, Lead.id == Appointment.lead_id
            ).filter(
                Lead.user_id == user.id,
                Lead.created_at >= start_date,
                Lead.created_at <= end_date
            ).distinct().count()
            
            user_conversion_rate = 0
            if user_leads > 0:
                user_conversion_rate = (user_conversions / user_leads) * 100
            
            # Calculate average response time for this user
            user_response_times = []
            user_leads_with_comms = db.session.query(
                Lead.id, Lead.created_at, func.min(Communication.sent_at).label('first_comm')
            ).join(
                Communication, Lead.id == Communication.lead_id
            ).filter(
                Lead.user_id == user.id,
                Lead.created_at >= start_date,
                Lead.created_at <= end_date
            ).group_by(Lead.id).all()
            
            for lead in user_leads_with_comms:
                if lead.created_at and lead.first_comm:
                    time_diff = (lead.first_comm - lead.created_at).total_seconds() / 60
                    user_response_times.append(time_diff)
            
            user_avg_response_time = 0
            if user_response_times:
                user_avg_response_time = round(sum(user_response_times) / len(user_response_times), 1)
            
            user_performance.append({
                'Username': user.username,
                'Leads': user_leads,
                'Appointments': user_appointments,
                'Kept Appointments': user_kept_appointments,
                'Conversions': user_conversions,
                'Conversion Rate (%)': round(user_conversion_rate, 1),
                'Avg. Response Time (min)': user_avg_response_time
            })
        
        # Export data based on format
        if format == 'csv':
            # Create CSV
            import csv
            fieldnames = ['Username', 'Leads', 'Appointments', 'Kept Appointments', 
                         'Conversions', 'Conversion Rate (%)', 'Avg. Response Time (min)']
            
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for user_data in user_performance:
                writer.writerow(user_data)
            
            output.seek(0)
            
            # Create response
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=performance_report_{start_date}_to_{end_date}.csv'
            response.headers['Content-type'] = 'text/csv'
            
            return response
            
        elif format == 'excel':
            # Create Excel file
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Convert to DataFrame
            df = pd.DataFrame(user_performance)
            
            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Performance Metrics"
            
            # Add title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = f"Performance Metrics Report ({start_date} to {end_date})"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add headers
            headers = list(user_performance[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                
                # Add border
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
            
            # Add data
            for row_num, user_data in enumerate(user_performance, 4):
                for col_num, (key, value) in enumerate(user_data.items(), 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Add border
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to BytesIO
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=performance_report_{start_date}_to_{end_date}.xlsx'
            response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return response
    
    elif report_type == 'leads':
        # Get lead data
        leads = Lead.query.filter(
            Lead.created_at >= start_date,
            Lead.created_at <= end_date
        ).all()
        
        lead_data = []
        for lead in leads:
            lead_data.append({
                'Lead ID': lead.id,
                'Name': f"{lead.first_name} {lead.last_name}",
                'Email': lead.email,
                'Phone': lead.phone,
                'Source': lead.source,
                'Status': lead.status,
                'Created At': lead.created_at.strftime('%Y-%m-%d %H:%M'),
                'Assigned To': User.query.get(lead.user_id).username if lead.user_id else 'Unassigned'
            })
        
        # Export data based on format
        if format == 'csv':
            import csv
            fieldnames = ['Lead ID', 'Name', 'Email', 'Phone', 'Source', 'Status', 'Created At', 'Assigned To']
            
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for lead in lead_data:
                writer.writerow(lead)
            
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=leads_report_{start_date}_to_{end_date}.csv'
            response.headers['Content-type'] = 'text/csv'
            
            return response
            
        elif format == 'excel':
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Convert to DataFrame
            df = pd.DataFrame(lead_data)
            
            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads Report"
            
            # Add title
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"Leads Report ({start_date} to {end_date})"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add headers
            headers = list(lead_data[0].keys()) if lead_data else []
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            # Add data
            for row_num, lead in enumerate(lead_data, 4):
                for col_num, (key, value) in enumerate(lead.items(), 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to BytesIO
            wb.save(output)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=leads_report_{start_date}_to_{end_date}.xlsx'
            response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return response
    
    elif report_type == 'appointments':
        # Get appointment data
        appointments = Appointment.query.filter(
            Appointment.date >= start_date,
            Appointment.date <= end_date
        ).all()
        
        appointment_data = []
        for appointment in appointments:
            lead = Lead.query.get(appointment.lead_id)
            appointment_data.append({
                'Appointment ID': appointment.id,
                'Lead': f"{lead.first_name} {lead.last_name}" if lead else 'Unknown',
                'Date': appointment.date.strftime('%Y-%m-%d'),
                'Time': appointment.time.strftime('%H:%M') if appointment.time else 'N/A',
                'Purpose': appointment.purpose,
                'Status': appointment.status,
                'Scheduled By': User.query.get(appointment.user_id).username if appointment.user_id else 'System'
            })
        
        # Export data based on format
        if format == 'csv':
            import csv
            fieldnames = ['Appointment ID', 'Lead', 'Date', 'Time', 'Purpose', 'Status', 'Scheduled By']
            
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for appointment in appointment_data:
                writer.writerow(appointment)
            
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=appointments_report_{start_date}_to_{end_date}.csv'
            response.headers['Content-type'] = 'text/csv'
            
            return response
            
        elif format == 'excel':
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Convert to DataFrame
            df = pd.DataFrame(appointment_data)
            
            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Appointments Report"
            
            # Add title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = f"Appointments Report ({start_date} to {end_date})"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add headers
            headers = list(appointment_data[0].keys()) if appointment_data else []
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            # Add data
            for row_num, appointment in enumerate(appointment_data, 4):
                for col_num, (key, value) in enumerate(appointment.items(), 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to BytesIO
            wb.save(output)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=appointments_report_{start_date}_to_{end_date}.xlsx'
            response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return response
    
    # If we get here, redirect to dashboard with error
    flash('Invalid report type', 'danger')
    return redirect(url_for('analytics.dashboard'))

@bp.route('/api/chart_data/<chart_type>')
@login_required
def chart_data(chart_type):
    """API endpoint to get chart data for the analytics dashboards"""
    # Get date range from request parameters or use default (last 30 days)
    end_date = request.args.get('end_date', datetime.now().date().isoformat())
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).date().isoformat())
    
    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    except ValueError:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
    
    if chart_type == 'lead_sources':
        data = db.session.query(
            Lead.source, func.count(Lead.id).label('count')
        ).group_by(Lead.source).all()
        
        return jsonify({
            'labels': [item[0] or 'Unknown' for item in data],
            'data': [item[1] for item in data]
        })
        
    elif chart_type == 'lead_statuses':
        data = db.session.query(
            Lead.status, func.count(Lead.id).label('count')
        ).group_by(Lead.status).all()
        
        return jsonify({
            'labels': [item[0] for item in data],
            'data': [item[1] for item in data]
        })
        
    elif chart_type == 'monthly_leads':
        year = datetime.now().year
        data = []
        
        for month in range(1, 13):
            count = Lead.query.filter(
                extract('year', Lead.created_at) == year,
                extract('month', Lead.created_at) == month
            ).count()
            data.append(count)
        
        return jsonify({
            'labels': [calendar.month_name[month] for month in range(1, 13)],
            'data': data
        })
        
    elif chart_type == 'appointment_statuses':
        data = db.session.query(
            Appointment.status, func.count(Appointment.id).label('count')
        ).group_by(Appointment.status).all()
        
        return jsonify({
            'labels': [item[0] for item in data],
            'data': [item[1] for item in data]
        })
        
    elif chart_type == 'communication_types':
        data = db.session.query(
            Communication.type, func.count(Communication.id).label('count')
        ).group_by(Communication.type).all()
        
        return jsonify({
            'labels': [item[0] for item in data],
            'data': [item[1] for item in data]
        })
        
    elif chart_type == 'user_performance':
        if current_user.role not in ['admin', 'manager']:
            return jsonify({'error': 'Unauthorized access'}), 403
            
        data = db.session.query(
            User.username,
            func.count(Lead.id).label('leads')
        ).outerjoin(Lead, User.id == Lead.user_id
        ).group_by(User.username).all()
        
        return jsonify({
            'labels': [item[0] for item in data],
            'data': [item[1] for item in data]
        })
        
    elif chart_type == 'conversion_funnel':
        # Conversion funnel data
        total_leads = Lead.query.filter(
            Lead.created_at >= start_date,
            Lead.created_at <= end_date
        ).count()
        
        contacted_leads = db.session.query(Lead.id).join(
            Communication, Lead.id == Communication.lead_id
        ).filter(
            Lead.created_at >= start_date,
            Lead.created_at <= end_date
        ).distinct().count()
        
        appointments_scheduled = db.session.query(Lead.id).join(
            Appointment, Lead.id == Appointment.lead_id
        ).filter(
            Lead.created_at >= start_date,
            Lead.created_at <= end_date
        ).distinct().count()
        
        appointments_kept = db.session.query(Lead.id).join(
            Appointment, Lead.id == Appointment.lead_id
        ).filter(
            Lead.created_at >= start_date,
            Lead.created_at <= end_date,
            Appointment.status == 'completed'
        ).distinct().count()
        
        converted_leads = Lead.query.filter(
            Lead.created_at >= start_date,
            Lead.created_at <= end_date,
            Lead.status == 'converted'
        ).count()
        
        labels = ['Total Leads', 'Contacted', 'Appointment Scheduled', 'Appointment Kept', 'Converted']
        values = [total_leads, contacted_leads, appointments_scheduled, appointments_kept, converted_leads]
        
        return jsonify({
            'labels': labels,
            'values': values
        })
    
    elif chart_type == 'lead_source_performance':
        # Lead source performance (conversion rate by source)
        sources = db.session.query(Lead.source).distinct().all()
        labels = []
        values = []
        
        for source in sources:
            source_name = source[0]
            total_from_source = Lead.query.filter(
                Lead.source == source_name,
                Lead.created_at >= start_date,
                Lead.created_at <= end_date
            ).count()
            
            if total_from_source == 0:
                continue
                
            converted_from_source = db.session.query(Lead.id).join(
                Appointment, Lead.id == Appointment.lead_id
            ).filter(
                Lead.source == source_name,
                Lead.created_at >= start_date,
                Lead.created_at <= end_date
            ).distinct().count()
            
            conversion_rate = (converted_from_source / total_from_source) * 100
            
            labels.append(source_name)
            values.append(round(conversion_rate, 1))
        
        return jsonify({
            'labels': labels,
            'values': values
        })
    
    elif chart_type == 'response_time_trend':
        # Response time trend (average by week)
        weeks = []
        current_date = start_date
        while current_date <= end_date:
            week_end = min(current_date + timedelta(days=6), end_date)
            weeks.append((current_date, week_end))
            current_date = week_end + timedelta(days=1)
        
        labels = []
        values = []
        
        for week_start, week_end in weeks:
            week_label = f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d')}"
            
            leads_with_comms = db.session.query(
                Lead.id, Lead.created_at, func.min(Communication.sent_at).label('first_comm')
            ).join(
                Communication, Lead.id == Communication.lead_id
            ).filter(
                Lead.created_at >= week_start,
                Lead.created_at <= week_end
            ).group_by(Lead.id).all()
            
            response_times = []
            for lead in leads_with_comms:
                if lead.created_at and lead.first_comm:
                    time_diff = (lead.first_comm - lead.created_at).total_seconds() / 60
                    response_times.append(time_diff)
            
            avg_time = 0
            if response_times:
                avg_time = round(sum(response_times) / len(response_times), 1)
            
            labels.append(week_label)
            values.append(avg_time)
        
        return jsonify({
            'labels': labels,
            'values': values
        })
    
    elif chart_type == 'appointment_outcomes':
        # Appointment outcomes
        appointment_outcomes = db.session.query(
            Appointment.status, func.count(Appointment.id)
        ).filter(
            Appointment.date >= start_date,
            Appointment.date <= end_date
        ).group_by(Appointment.status).all()
        
        labels = [status for status, count in appointment_outcomes]
        values = [count for status, count in appointment_outcomes]
        
        return jsonify({
            'labels': labels,
            'values': values
        })
    
    elif chart_type == 'team_performance':
        # Team performance by selected metric
        metric = request.args.get('metric', 'leads')
        users = User.query.all()
        labels = []
        values = []
        metric_label = ''
        
        for user in users:
            labels.append(user.username)
            
            if metric == 'leads':
                metric_label = 'Leads'
                count = Lead.query.filter(
                    Lead.user_id == user.id,
                    Lead.created_at >= start_date,
                    Lead.created_at <= end_date
                ).count()
                values.append(count)
                
            elif metric == 'appointments':
                metric_label = 'Appointments'
                count = Appointment.query.filter(
                    Appointment.user_id == user.id,
                    Appointment.date >= start_date,
                    Appointment.date <= end_date
                ).count()
                values.append(count)
                
            elif metric == 'conversions':
                metric_label = 'Conversion Rate (%)'
                user_leads = Lead.query.filter(
                    Lead.user_id == user.id,
                    Lead.created_at >= start_date,
                    Lead.created_at <= end_date
                ).count()
                
                user_conversions = db.session.query(Lead.id).join(
                    Appointment, Lead.id == Appointment.lead_id
                ).filter(
                    Lead.user_id == user.id,
                    Lead.created_at >= start_date,
                    Lead.created_at <= end_date
                ).distinct().count()
                
                conversion_rate = 0
                if user_leads > 0:
                    conversion_rate = (user_conversions / user_leads) * 100
                
                values.append(round(conversion_rate, 1))
                
            elif metric == 'response_time':
                metric_label = 'Avg. Response Time (min)'
                user_response_times = []
                user_leads_with_comms = db.session.query(
                    Lead.id, Lead.created_at, func.min(Communication.sent_at).label('first_comm')
                ).join(
                    Communication, Lead.id == Communication.lead_id
                ).filter(
                    Lead.user_id == user.id,
                    Lead.created_at >= start_date,
                    Lead.created_at <= end_date
                ).group_by(Lead.id).all()
                
                for lead in user_leads_with_comms:
                    if lead.created_at and lead.first_comm:
                        time_diff = (lead.first_comm - lead.created_at).total_seconds() / 60
                        user_response_times.append(time_diff)
                
                avg_time = 0
                if user_response_times:
                    avg_time = round(sum(user_response_times) / len(user_response_times), 1)
                
                values.append(avg_time)
        
        return jsonify({
            'labels': labels,
            'values': values,
            'metric_label': metric_label
        })
    
    # Default response if chart type not recognized
    return jsonify({
        'labels': [],
        'values': []
    })

@bp.route('/performance')
@login_required
@manager_required
def performance_metrics():
    """Performance metrics dashboard with KPIs and detailed analytics"""
    # Date ranges for filtering
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    if request.args.get('start_date') and request.args.get('end_date'):
        try:
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Key Performance Indicators
    
    # 1. Lead Conversion Rate
    total_leads = Lead.query.filter(
        func.date(Lead.created_at) >= start_date,
        func.date(Lead.created_at) <= end_date
    ).count()
    
    converted_leads = db.session.query(Lead).filter(
        func.date(Lead.created_at) >= start_date,
        func.date(Lead.created_at) <= end_date,
        Lead.status.in_(['Qualified', 'Appointment Set', 'Sale Completed'])
    ).count()
    
    lead_conversion_rate = round((converted_leads / total_leads * 100), 1) if total_leads > 0 else 0
    
    # 2. Average Response Time
    response_times = []
    leads_with_comms = db.session.query(Lead).filter(
        func.date(Lead.created_at) >= start_date,
        func.date(Lead.created_at) <= end_date,
        Lead.communications.any()
    ).all()
    
    for lead in leads_with_comms:
        first_comm = lead.communications.order_by(Communication.sent_at).first()
        if first_comm:
            response_time = (first_comm.sent_at - lead.created_at).total_seconds() / 3600  # hours
            response_times.append(response_time)
    
    avg_response_time = round(sum(response_times) / len(response_times), 1) if response_times else 0
    
    # 3. Appointment Kept Rate
    total_appointments = Appointment.query.filter(
        func.date(Appointment.date) >= start_date,
        func.date(Appointment.date) <= end_date
    ).count()
    
    kept_appointments = Appointment.query.filter(
        func.date(Appointment.date) >= start_date,
        func.date(Appointment.date) <= end_date,
        Appointment.status == 'Completed'
    ).count()
    
    appointment_kept_rate = round((kept_appointments / total_appointments * 100), 1) if total_appointments > 0 else 0
    
    # 4. Cost Per Lead (assuming a fixed cost per lead for demonstration)
    # In a real application, this would be calculated based on actual marketing spend
    cost_per_lead = round(50.0, 2)  # Example fixed cost
    
    # User Performance Metrics
    user_performance = db.session.query(
        User.id,
        User.username,
        User.first_name,
        User.last_name,
        func.count(Lead.id).label('leads_count'),
        func.count(case([(Lead.status.in_(['Qualified', 'Appointment Set', 'Sale Completed']), 1)])).label('converted_leads'),
        func.count(Appointment.id).label('appointments_count'),
        func.count(case([(Appointment.status == 'Completed', 1)])).label('kept_appointments'),
        func.count(Communication.id).label('communications_count')
    ).outerjoin(Lead, User.id == Lead.user_id
    ).outerjoin(Appointment, User.id == Appointment.user_id
    ).outerjoin(Communication, User.id == Communication.user_id
    ).filter(User.role == 'bdc_agent'
    ).group_by(User.id, User.username, User.first_name, User.last_name).all()
    
    # Calculate derived metrics for each user
    user_metrics = []
    for user in user_performance:
        user_dict = {
            'id': user.id,
            'username': user.username,
            'name': f"{user.first_name} {user.last_name}" if user.first_name and user.last_name else user.username,
            'leads_count': user.leads_count,
            'converted_leads': user.converted_leads,
            'conversion_rate': round((user.converted_leads / user.leads_count * 100), 1) if user.leads_count > 0 else 0,
            'appointments_count': user.appointments_count,
            'kept_appointments': user.kept_appointments,
            'appointment_kept_rate': round((user.kept_appointments / user.appointments_count * 100), 1) if user.appointments_count > 0 else 0,
            'communications_count': user.communications_count,
            'avg_comms_per_lead': round((user.communications_count / user.leads_count), 1) if user.leads_count > 0 else 0
        }
        user_metrics.append(user_dict)
    
    # Sort users by lead count (descending)
    user_metrics.sort(key=lambda x: x['leads_count'], reverse=True)
    
    return render_template('analytics/performance.html',
                          title='Performance Metrics',
                          start_date=start_date,
                          end_date=end_date,
                          total_leads=total_leads,
                          converted_leads=converted_leads,
                          lead_conversion_rate=lead_conversion_rate,
                          avg_response_time=avg_response_time,
                          total_appointments=total_appointments,
                          kept_appointments=kept_appointments,
                          appointment_kept_rate=appointment_kept_rate,
                          cost_per_lead=cost_per_lead,
                          user_metrics=user_metrics)

@bp.route('/campaigns')
@login_required
@manager_required
def campaign_analytics():
    """Marketing campaign analytics dashboard"""
    # Date ranges for filtering
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    if request.args.get('start_date') and request.args.get('end_date'):
        try:
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get all campaigns
    from app.models import MarketingCampaign, CampaignChannel, CampaignMetric
    
    # Active campaigns count
    active_campaigns = MarketingCampaign.query.filter(
        MarketingCampaign.status == 'active'
    ).count()
    
    # Get all campaigns for the detail section
    campaigns = MarketingCampaign.query.filter(
        or_(
            and_(MarketingCampaign.start_date >= start_date, MarketingCampaign.start_date <= end_date),
            and_(MarketingCampaign.end_date >= start_date, MarketingCampaign.end_date <= end_date),
            and_(MarketingCampaign.start_date <= start_date, or_(MarketingCampaign.end_date >= end_date, MarketingCampaign.end_date == None))
        )
    ).all()
    
    # Calculate total leads from all campaigns
    total_leads = sum(campaign.total_leads() for campaign in campaigns)
    
    # Calculate average cost per lead across all campaigns
    campaigns_with_budget = [c for c in campaigns if c.budget is not None and c.total_leads() > 0]
    if campaigns_with_budget:
        avg_cost_per_lead = round(sum(c.cost_per_lead() for c in campaigns_with_budget) / len(campaigns_with_budget), 2)
    else:
        avg_cost_per_lead = 0
    
    # Calculate average conversion rate across all campaigns
    campaigns_with_leads = [c for c in campaigns if c.total_leads() > 0]
    if campaigns_with_leads:
        avg_conversion_rate = round(sum(c.conversion_rate() for c in campaigns_with_leads) / len(campaigns_with_leads), 1)
    else:
        avg_conversion_rate = 0
    
    # Add budget spent calculation for each campaign (in a real app, this would come from actual data)
    for campaign in campaigns:
        # For demo purposes, calculate a random percentage of budget spent
        import random
        if campaign.budget:
            campaign.budget_spent = round(campaign.budget * random.uniform(0.1, 0.9), 2)
        else:
            campaign.budget_spent = 0
    
    return render_template('analytics/campaign_analytics.html',
                          title='Campaign Analytics',
                          start_date=start_date,
                          end_date=end_date,
                          active_campaigns=active_campaigns,
                          total_leads=total_leads,
                          avg_cost_per_lead=avg_cost_per_lead,
                          avg_conversion_rate=avg_conversion_rate,
                          campaigns=campaigns)

@bp.route('/campaigns/<int:campaign_id>')
@login_required
@manager_required
def campaign_detail(campaign_id):
    """Detailed view of a specific marketing campaign"""
    from app.models import MarketingCampaign, CampaignChannel, CampaignMetric, Lead
    
    campaign = MarketingCampaign.query.get_or_404(campaign_id)
    
    # Get campaign metrics over time
    metrics = CampaignMetric.query.filter(
        CampaignMetric.campaign_id == campaign_id
    ).order_by(CampaignMetric.date).all()
    
    # Group metrics by date and type
    metrics_by_date = {}
    for metric in metrics:
        date_str = metric.date.strftime('%Y-%m-%d')
        if date_str not in metrics_by_date:
            metrics_by_date[date_str] = {}
        metrics_by_date[date_str][metric.metric_name] = metric.metric_value
    
    # Get leads from this campaign
    campaign_leads = Lead.query.join(
        'campaigns'
    ).filter(
        MarketingCampaign.id == campaign_id
    ).all()
    
    # Calculate lead status distribution
    lead_statuses = {}
    for lead in campaign_leads:
        if lead.status not in lead_statuses:
            lead_statuses[lead.status] = 0
        lead_statuses[lead.status] += 1
    
    # Calculate lead source distribution
    lead_sources = {}
    for lead in campaign_leads:
        if lead.source not in lead_sources:
            lead_sources[lead.source] = 0
        lead_sources[lead.source] += 1
    
    return render_template('analytics/campaign_detail.html',
                          title=f'Campaign: {campaign.name}',
                          campaign=campaign,
                          metrics_by_date=metrics_by_date,
                          lead_statuses=lead_statuses,
                          lead_sources=lead_sources,
                          campaign_leads=campaign_leads)
